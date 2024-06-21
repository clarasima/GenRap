from flask import abort, Response
from app import publications_collection, users_collection, citations_collection
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.writer.excel import save_virtual_workbook

def validate_request_data(data):
    if not data or 'ids' not in data or 'type' not in data or 'fields' not in data:
        abort(400, description='IDs, type, or fields missing')
    return data['ids'], data['type'], data['fields']

def get_collection_and_filename(doc_type):
    if doc_type == 'publications':
        return publications_collection, 'publications', 'Publication'
    elif doc_type == 'citations':
        return citations_collection, 'citations', 'Citation'
    else:
        abort(400, description='Invalid type specified')

def retrieve_documents(collection, object_ids):
    documents_cursor = collection.find({"_id": {"$in": object_ids}})
    documents = list(documents_cursor)
    if not documents:
        abort(404, description=f'No documents found for the provided IDs')
    return documents

def update_user_collection(documents, selected_fields):
    user_id = documents[0].get('user_id')
    if user_id:
        users_collection.update_one(
            {'_id': user_id},
            {'$set': {'desired_fields': selected_fields}},
            upsert=True
        )

def sort_documents(documents, selected_fields):
    if 'year' in selected_fields:
        documents.sort(key=lambda x: x.get('edited', {}).get('year', float('-inf')), reverse=True)
    return documents

def create_txt_response(documents, selected_fields, filename, doc_name):
    output = io.StringIO()
    for i, doc in enumerate(documents, start=1):
        edited = doc.get('edited', {})
        addedFields = edited.get("addedFields", {})
        output.write(f'{doc_name} #{i}\n')
        for field in selected_fields:
            value = edited.get(field) or addedFields.get(field) or '-'
            if isinstance(value, list):
                value = ", ".join(value)
            output.write(f'{field.capitalize().replace("_", " ")}: {value}\n')
        output.write('\n\n')
    response = Response(output.getvalue(), mimetype='text/plain')
    response.headers.set('Content-Disposition', f'attachment; filename={filename}.txt')
    return response

def create_excel_response(documents, selected_fields, filename, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    headers = [field.capitalize().replace("_", " ") for field in selected_fields]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    for doc in documents:
        edited = doc.get('edited', {})
        addedFields = edited.get("addedFields", {})
        row = [edited.get(field) or addedFields.get(field) or '-' for field in selected_fields]
        if 'authors' in selected_fields:
            if isinstance(row[selected_fields.index('authors')], list):
                row[selected_fields.index('authors')] = ', '.join(row[selected_fields.index('authors')])
            
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    response = Response(save_virtual_workbook(wb), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers.set('Content-Disposition', f'attachment; filename={filename}.xlsx')
    return response

def create_csv_response(documents, selected_fields, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    headers = [field.capitalize().replace("_", " ") for field in selected_fields]
    writer.writerow(headers)
    for doc in documents:
        edited = doc.get('edited', {})
        addedFields = edited.get("addedFields", {})
        row = [edited.get(field) or addedFields.get(field) or '-' for field in selected_fields]
        if 'authors' in selected_fields:
            if isinstance(row[selected_fields.index('authors')], list):
                row[selected_fields.index('authors')] = ', '.join(row[selected_fields.index('authors')])
        writer.writerow(row)
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers.set('Content-Disposition', f'attachment; filename={filename}.csv')
    return response